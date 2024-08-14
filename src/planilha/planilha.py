import openpyxl

def get_number_for_letter(letter):
    normalized_letter = letter.upper()
    return ord(normalized_letter) - 64

mesesstr = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
mesesInt = {mes: i + 1 for i, mes in enumerate(mesesstr)}
class Planilha:
    
    linha_faturamento = 14
    
    def __init__(self, arquivo):
        self.arquivo = arquivo
        self.workbook = openpyxl.load_workbook(arquivo)
        
        self.linha_faturamento = self.get_linha_com_descricao_aba_apresentacao('FATURAMENTO')

        
    def set_CNPJ(self, cnpj):
        self.workbook['Apresentação']['C7'] = 'CNPJ: ' + cnpj
        
    def set_EMPRESA(self, empresa):
        self.workbook['Apresentação']['B7'] = 'EMPRESA: ' + empresa
    
    def inserir_colunas_mes_aba_dados(self, mesInicio: int, anoInicio: int, mesFim: int, anoFim: int):
        aba_dados = self.workbook['Dados']
        
        coluna_inicio = get_number_for_letter('B')        
        
        for ano in range(anoInicio, anoFim + 1):
            for mes in range(mesInicio, mesFim + 1):
                coluna = coluna_inicio + (ano - anoInicio) * 12 + mes - mesInicio
                aba_dados.insert_cols(coluna)
                
                messtr = mesesstr[mes - 1]
                aba_dados.cell(row=1, column=coluna, value=f'{messtr}/{ano}')
                
        coluna_total = coluna + 1
        aba_dados.insert_cols(coluna_total)
        aba_dados.cell(row=1, column=coluna_total, value='Total')
        
        self.ajustar_width_colunas_aba('Dados')
                
        def inserir_periodo(mesInicio, anoInicio, mesFim, anoFim):
            
            
            coluna_periodo = get_number_for_letter('C')
            linha_periodo = 10
            
            periodo = f'Período: {mesesstr[mesInicio - 1]}/{anoInicio} a {mesesstr[mesFim - 1]}/{anoFim}'
            
            self.workbook['Apresentação'].cell(row=linha_periodo, column=coluna_periodo, value=periodo)            
        
        inserir_periodo(mesInicio, anoInicio, mesFim, anoFim)
    
    def ajustar_width_colunas_aba(self, nome_aba):
        aba = self.workbook[nome_aba]
        
        for col in aba.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            aba.column_dimensions[column].width = adjusted_width

    def insert_dados_aba_dados(self, lista_de_dados: list[dict[str, str | float]], is_insert_dados_aba_apresentacao: bool = False):  
        '''
            lista_de_dados: lista de dicionários onde cada dicionário representa um dado a ser inserido na planilha. Exemplo:
                [
                    {
                        'descricao': 'Receita',
                        '2019-01': 100,
                        '2019-02': 200
                    },
                    {
                        'descricao': 'Despesa',
                        '2019-01': 50,
                        '2019-02': 100
                    }
                ]
            is_insert_dados_aba_apresentacao: se True, insere a linha na aba de apresentação acima de TRIBUTOS SOBRE VENDAS
        '''
        
        aba_dados = self.workbook['Dados']
        
        ultima_linha_preenchida = aba_dados.max_row
        
        linha = ultima_linha_preenchida
        for dado in lista_de_dados:
            linha += 1
            
            aba_dados.cell(row=linha, column=1, value=dado['descricao'])
            
            ultima_coluna = aba_dados.max_column
            
            for coluna in range(2, ultima_coluna + 1):
                header = aba_dados.cell(row=1, column=coluna).value
                is_coluna_total = header == 'Total'
                if is_coluna_total:
                    formula_sum = f'=SUM({openpyxl.utils.get_column_letter(2)}{linha}:{openpyxl.utils.get_column_letter(ultima_coluna - 1)}{linha})'
                    aba_dados.cell(row=linha, column=coluna, value=formula_sum)
                    break
                
                mes, ano = header.split('/')
                mes = mesesInt[mes]
                    
                mes_MM = str(mes).zfill(2)
                valor_mes = dado[f'{ano}-{mes_MM}'] if f'{ano}-{mes_MM}' in dado else None
                aba_dados.cell(row=linha, column=coluna, value=valor_mes)
            
            if is_insert_dados_aba_apresentacao:
                self.inserir_linha_dados_na_apresentacao(linha)
    
    def get_linha_com_descricao_aba_apresentacao(self, descricao: str):
        aba_apresentacao = self.workbook['Apresentação']
        
        linha_descricao = None
        ultima_linha = aba_apresentacao.max_row
        for linha in range(1, ultima_linha + 1):
            if descricao == aba_apresentacao.cell(row=linha, column=2).value:
                linha_descricao = linha
                break
        
        return linha_descricao
    
    
    def get_linha_com_descricao_aba_dados(self, descricao: str):
        aba_dados = self.workbook['Dados']
        
        linha_descricao = None
        ultima_linha = aba_dados.max_row
        for linha in range(1, ultima_linha + 1):
            if descricao == aba_dados.cell(row=linha, column=1).value:
                linha_descricao = linha
                break
        
        return linha_descricao
    
    def inserir_valor_na_linha_aba_apresentacao(self, linha: int, valor: str, quantidade: str | int, quantidade_formato: str = '0'):
        aba_apresentacao = self.workbook['Apresentação']
        
        celula_faturamento = "$F$" + str(self.linha_faturamento)
        formula_porcentagem_total_sobre_faturamento = f'=F{linha}/{celula_faturamento}'
                
        coluna_F = get_number_for_letter('F')
        coluna_G = get_number_for_letter('G')
        coluna_H = get_number_for_letter('H')        
        
        aba_apresentacao.cell(row=linha, column=coluna_F, value=valor).number_format = '#,##0.00'
        aba_apresentacao.cell(row=linha, column=coluna_F).alignment = openpyxl.styles.Alignment(horizontal='center')
        
        aba_apresentacao.cell(row=linha, column=coluna_G, value=formula_porcentagem_total_sobre_faturamento).number_format = '0.00%'
        aba_apresentacao.cell(row=linha, column=coluna_G).alignment = openpyxl.styles.Alignment(horizontal='center')
        
        aba_apresentacao.cell(row=linha, column=coluna_H, value=quantidade).number_format = quantidade_formato
        aba_apresentacao.cell(row=linha, column=coluna_H).alignment = openpyxl.styles.Alignment(horizontal='center')
    
    def get_dado_by_linha(self, linha: int) -> dict:
        aba_dados = self.workbook['Dados']
        
        descricao = aba_dados.cell(row=linha, column=1).value
        formula_total = f'=Dados!{openpyxl.utils.get_column_letter(aba_dados.max_column)}{linha}'
        formula_quantidade_meses_com_valor = f'=COUNTIF(Dados!B{linha}:{openpyxl.utils.get_column_letter(aba_dados.max_column - 1)}{linha}, "<>")'
        
        return {
            'descricao': descricao,
            'formula_total': formula_total,
            'formula_quantidade_meses_com_valor': formula_quantidade_meses_com_valor
        }
    
    def inserir_linha_aba_apresentacao(self, descricao: str, descricao_linha_para_inserir_acima: str = '% TRIBUTOS SOBRE VENDAS') -> int:
        linha_para_inserir_acima = self.get_linha_com_descricao_aba_apresentacao(descricao_linha_para_inserir_acima)
        
        #insere linha acima da linha de tributos sobre vendas
        if linha_para_inserir_acima is None:
            return
        
        aba_apresentacao = self.workbook['Apresentação']
        aba_apresentacao.insert_rows(linha_para_inserir_acima)
        
        #borda espessa esquerda na "B" e borda espessa direita ta "H"
        coluna_B = get_number_for_letter('B')
        coluna_H = get_number_for_letter('H')
        
        aba_apresentacao.cell(row=linha_para_inserir_acima, column=coluna_B).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='medium'))
        aba_apresentacao.cell(row=linha_para_inserir_acima, column=coluna_H).border = openpyxl.styles.Border(right=openpyxl.styles.Side(style='medium'))
        
        aba_apresentacao.cell(row=linha_para_inserir_acima, column=coluna_B, value=descricao)
        
        linha_para_inserir = linha_para_inserir_acima
        return linha_para_inserir
    
    def inserir_linha_dados_na_apresentacao(self, linha: int):
        dado = self.get_dado_by_linha(linha)
        descricao = dado['descricao']
        formula_total = dado['formula_total']
        formula_quantidade_meses_com_valor = dado['formula_quantidade_meses_com_valor']
            
        linha_para_inserir = self.inserir_linha_aba_apresentacao(descricao)
        
        self.inserir_valor_na_linha_aba_apresentacao(linha_para_inserir, formula_total, formula_quantidade_meses_com_valor)                
            
            
    def inserir_valor_dado_na_apresentacao_pela_descricao(self, descricao_apresentacao, descricao_dados):
        linha_dados = self.get_linha_com_descricao_aba_dados(descricao_dados)
        if linha_dados is None:
            return
        
        linha_apresentacao = self.get_linha_com_descricao_aba_apresentacao(descricao_apresentacao)
        if linha_apresentacao is None:
            return
        
        dado = self.get_dado_by_linha(linha_dados)
        formula_total = dado['formula_total']
        formula_quantidade_meses_com_valor = dado['formula_quantidade_meses_com_valor']
        
        self.inserir_valor_na_linha_aba_apresentacao(linha_apresentacao, formula_total, formula_quantidade_meses_com_valor)
        
    def get_formula_quantidade_meses_com_valor_e_descricao(self, descricao: str):
        aba_dados = self.workbook['Dados']
        
        coluna_total = aba_dados.max_column
        coluna_maxima = openpyxl.utils.get_column_letter(coluna_total - 2)
        
        ultima_linha = aba_dados.max_row
        
        formula_quantidade_meses_com_valor = f'=SUMPRODUCT((ISNUMBER(SEARCH("{descricao}", Dados!A2:A{ultima_linha})))*(Dados!B2:{coluna_maxima}{ultima_linha} <> ""))'
        
        return formula_quantidade_meses_com_valor
        
    def inserir_soma_dados_na_apresentacao_por_regex(self, descricao_contains: str, descricao_apresentacao: str):
        aba_dados = self.workbook['Dados']
        
        linha_apresentacao = self.get_linha_com_descricao_aba_apresentacao(descricao_apresentacao)
        if linha_apresentacao is None:
            return
        
        coluna_total = aba_dados.max_column
        
        formula_total = f'=SUMIF(Dados!A:A, "*{descricao_contains}*", Dados!{openpyxl.utils.get_column_letter(coluna_total)}:{openpyxl.utils.get_column_letter(coluna_total)})'
        formula_quantidade_meses_com_valor = self.get_formula_quantidade_meses_com_valor_e_descricao(descricao_contains)
        
        self.inserir_valor_na_linha_aba_apresentacao(linha_apresentacao, formula_total, formula_quantidade_meses_com_valor)
    
    def inserir_soma_dados_na_apresentacao_por_regex_acima_de_TRIBUTOS(self, descricao_contains: str, descricao_apresentacao: str):
        # com base na descricao da aba dados, insere a soma dos valores na aba apresentacao acima de TRIBUTOS SOBRE VENDAS
        # Semelhante a inserir_soma_dados_na_apresentacao_por_regex, mas insere acima de TRIBUTOS SOBRE VENDAS e com uma descrição definida ao invés de uma que já esteja na planilha
        linha_para_inserir = self.inserir_linha_aba_apresentacao(descricao_apresentacao)
        
        if linha_para_inserir is None:
            return
        
        aba_dados = self.workbook['Dados']
        
        coluna_total = aba_dados.max_column
        
        formula_total = f'=SUMIF(Dados!A:A, "*{descricao_contains}*", Dados!{openpyxl.utils.get_column_letter(coluna_total)}:{openpyxl.utils.get_column_letter(coluna_total)})'
        formula_quantidade_meses_com_valor = self.get_formula_quantidade_meses_com_valor_e_descricao(descricao_contains)
        
        self.inserir_valor_na_linha_aba_apresentacao(linha_para_inserir, formula_total, formula_quantidade_meses_com_valor)
    
    def inserir_soma_dados_na_apresentacao_por_regex_acima_de_X(self, descricao_contains: str, descricao_apresentacao: str, descricao_x: str):
        # com base na descricao da aba dados, insere a soma dos valores na aba apresentacao acima de MARGENS DE LUCRO E CUSTO
        # Semelhante a inserir_soma_dados_na_apresentacao_por_regex, mas insere acima de MARGENS DE LUCRO E CUSTO e com uma descrição definida ao invés de uma que já esteja na planilha
        linha_para_inserir = self.inserir_linha_aba_apresentacao(descricao_apresentacao, descricao_linha_para_inserir_acima=descricao_x)
        
        if linha_para_inserir is None:
            return
        
        aba_dados = self.workbook['Dados']
        
        coluna_total = aba_dados.max_column
        
        formula_total = f'=SUMIF(Dados!A:A, "*{descricao_contains}*", Dados!{openpyxl.utils.get_column_letter(coluna_total)}:{openpyxl.utils.get_column_letter(coluna_total)})'
        formula_quantidade_meses_com_valor = self.get_formula_quantidade_meses_com_valor_e_descricao(descricao_contains)
        
        self.inserir_valor_na_linha_aba_apresentacao(linha_para_inserir, formula_total, formula_quantidade_meses_com_valor)
       
    def atualizar_custo_fixo(self):
        linha_custo_fixo = self.get_linha_com_descricao_aba_apresentacao('Custo Fixo - Teórico')
        
        formula_custo_fixo = f'=F{self.linha_faturamento} * H{linha_custo_fixo}'
        
        self.inserir_valor_na_linha_aba_apresentacao(linha_custo_fixo, formula_custo_fixo, 0, '0.00%')
    
    def atualizar_tributos_sobre_vendas(self):
        linha_tributos_sobre_vendas = self.get_linha_com_descricao_aba_apresentacao('% TRIBUTOS SOBRE VENDAS')
        
        linha_inicial_tributos_aba_apresentacao = self.linha_faturamento - 1
        linha_final_tributos_aba_apresentacao = linha_tributos_sobre_vendas - 1
        
        coluna_G = get_number_for_letter('G')
        formula_tributos_sobre_vendas = f'=SUMPRODUCT(SUMIF( B{linha_inicial_tributos_aba_apresentacao}:B{linha_final_tributos_aba_apresentacao}, \'Tributos sobre vendas\'!A1:A40, G{linha_inicial_tributos_aba_apresentacao}:G{linha_final_tributos_aba_apresentacao}))'
        
        aba_apresentacao = self.workbook['Apresentação']
        
        aba_apresentacao.cell(row=linha_tributos_sobre_vendas, column=coluna_G, value=formula_tributos_sobre_vendas).number_format = '0.00%'
        
    def atualizar_lucro_teorico(self):
        linha_compras = self.get_linha_com_descricao_aba_apresentacao('COMPRAS')
        linha_tributos_sobre_vendas = self.get_linha_com_descricao_aba_apresentacao('% TRIBUTOS SOBRE VENDAS')
        linha_custo_fixo = self.get_linha_com_descricao_aba_apresentacao('Custo Fixo - Teórico')
        linha_lucro_teorico = self.get_linha_com_descricao_aba_apresentacao('Lucro - Teórico')
        if linha_lucro_teorico is None:
            print('Não foi possível inserir linha de Lucro - Teórico')
            return              
         
        formula_lucro_teorico = f'=1 - SUM(G{linha_compras}:G{linha_tributos_sobre_vendas}, G{linha_custo_fixo})'
        formula_faturamento_vezes_lucro_teorico = f'=F{self.linha_faturamento} * G{linha_lucro_teorico}'
                
        coluna_G = get_number_for_letter('G')
        coluna_F = get_number_for_letter('F')
        
        aba_apresentacao = self.workbook['Apresentação']
        aba_apresentacao.cell(row=linha_lucro_teorico, column=coluna_G, value=formula_lucro_teorico).number_format = '0.00%'
        aba_apresentacao.cell(row=linha_lucro_teorico, column=coluna_F, value=formula_faturamento_vezes_lucro_teorico).number_format = '#,##0.00'
        
    def atualizar_total_carga_tributaria(self):
        linha_tributos_sobre_vendas = self.get_linha_com_descricao_aba_apresentacao('% TRIBUTOS SOBRE VENDAS')
        linha_margens_de_lucro_e_custo = self.get_linha_com_descricao_aba_apresentacao('MARGENS DE LUCRO E CUSTO')
        linha_total_carga_tributaria = self.get_linha_com_descricao_aba_apresentacao('TOTAL CARGA TRIBUTÁRIA')
        
        formula_total_carga_tributaria = f'=SUM(G{linha_tributos_sobre_vendas}:G{linha_margens_de_lucro_e_custo - 1})'
        
        coluna_G = get_number_for_letter('G')
        
        aba_apresentacao = self.workbook['Apresentação']
        aba_apresentacao.cell(row=linha_total_carga_tributaria, column=coluna_G, value=formula_total_carga_tributaria).number_format = '0.00%'
        
    def atualizar_tributos_sobre_faturamento(self):
        def atualizar_inss():
            linha_inss_1410 = self.get_linha_com_descricao_aba_apresentacao('1410 - INSS')
            linha_inss_tributos_sobre_faturamento = self.get_linha_com_descricao_aba_apresentacao('INSS')
            
            valor_inss_1410 = f'=F{linha_inss_1410}' if linha_inss_tributos_sobre_faturamento is not None else '0'
            quantidade_inss_1410 = f'=H{linha_inss_1410}' if linha_inss_1410 is not None else '0'
            
            coluna_F = get_number_for_letter('F')
            coluna_G = get_number_for_letter('G')
            coluna_H = get_number_for_letter('H')
            
            aba_apresentacao = self.workbook['Apresentação']
            aba_apresentacao.cell(row=linha_inss_tributos_sobre_faturamento, column=coluna_F, value=valor_inss_1410).number_format = '#,##0.00'
            aba_apresentacao.cell(row=linha_inss_tributos_sobre_faturamento, column=coluna_G, value=f'=F{linha_inss_tributos_sobre_faturamento}/F{self.linha_faturamento}').number_format = '0.00%'
            aba_apresentacao.cell(row=linha_inss_tributos_sobre_faturamento, column=coluna_H, value=quantidade_inss_1410).number_format = '0.00%'
            
        def atualizar_irpj():
            #[ ] IRPJ - se lucro teorico > 0  -> lucro * % da linha
            linha_irpj = self.get_linha_com_descricao_aba_apresentacao('IRPJ (não identificamos baixa, quebras, roubo)')
            linha_lucro_teorico = self.get_linha_com_descricao_aba_apresentacao('Lucro - Teórico')
            
            formula_irpj = f'=IF(F{linha_lucro_teorico} > 0, F{linha_lucro_teorico} * H{linha_irpj}, 0)'
            
            coluna_F = get_number_for_letter('F')
            coluna_G = get_number_for_letter('G')
            coluna_H = get_number_for_letter('H')
            
            aba_apresentacao = self.workbook['Apresentação']
            aba_apresentacao.cell(row=linha_irpj, column=coluna_F, value=formula_irpj).number_format = '#,##0.00'
            aba_apresentacao.cell(row=linha_irpj, column=coluna_G, value=f'=F{linha_irpj}/F{self.linha_faturamento}').number_format = '0.00%'
            
        
        atualizar_inss()
        atualizar_irpj()
        
        
    def save(self, output_path = 'output.xlsx'):
        self.ajustar_width_colunas_aba('Dados')
        self.atualizar_custo_fixo()
        self.atualizar_tributos_sobre_vendas()
        self.atualizar_lucro_teorico()
        self.atualizar_total_carga_tributaria()
        self.atualizar_tributos_sobre_faturamento()
        
        self.workbook.save(output_path)   
                                   
if __name__ == '__main__':
    planilha_path = 'template.xlsx'
    
    planilha = Planilha(planilha_path)
    
    planilha.set_CNPJ('123456789')
    planilha.set_EMPRESA('Empresa ABC')
    planilha.inserir_colunas_mes_aba_dados(1, 2019, 12, 2022)
    planilha.insert_dados_aba_dados([
        {
            'descricao': 'Receita',
            '2019-01': 100,
            '2019-02': 200,
        },
        {
            'descricao': 'Despesa',
            '2019-01': 50,
            '2019-02': 100,
        }
    ])
    planilha.inserir_soma_dados_na_apresentacao_por_regex_acima_de_X(
        descricao_contains='Receita',
        descricao_apresentacao='RECEITA',
        descricao_x='MARGENS DE LUCRO E CUSTO'
    )
    planilha.ajustar_width_colunas_aba('Dados')
    
    planilha.save('output.xlsx')